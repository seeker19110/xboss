#!/usr/bin/env python
"""Kịch bản E2E với HẠ TẦNG THẬT: API + Celery worker + Redis đang chạy.

Khác với `tests/test_e2e_takeoff.py` (chạy trong CI, thay broker bằng gọi đồng bộ), file
này không giả lập gì cả: nó gửi file lên API thật qua mạng, chờ worker rời thật sự nhặt
task qua Redis, rồi tải file Excel về và kiểm nội dung. Đây là lớp duy nhất bắt được các
lỗi chỉ xuất hiện khi ghép hạ tầng — worker không kết nối được broker, thư mục volume sai
quyền, biến môi trường thiếu, API key chặn nhầm...

Cách chạy:

    docker compose up --build -d
    uv run python scripts/e2e_smoke.py

    # API ở máy khác / có bật khóa API:
    E2E_BASE_URL=http://10.0.0.5:8083 MEP_AGENTS_API_KEY=... uv run python scripts/e2e_smoke.py

Mã thoát: 0 = đạt, 1 = có bước hỏng (in rõ bước nào).
"""
from __future__ import annotations

import io
import os
import sys
import time
import urllib.error
import urllib.request

BASE_URL = os.environ.get("E2E_BASE_URL", "http://localhost:8083").rstrip("/")
API_KEY = os.environ.get("MEP_AGENTS_API_KEY", "").strip()
TIMEOUT_S = float(os.environ.get("E2E_TIMEOUT", "180"))

MAIN_PIPE_LEN = 12_000.0   # mm
XREF_PIPE_LEN = 8_000.0    # mm
WASTE_FACTOR = 1.05
EXPECTED_M = (MAIN_PIPE_LEN + XREF_PIPE_LEN) / 1000.0 * WASTE_FACTOR


def _step(msg: str) -> None:
    print(f"→ {msg}", flush=True)


def _fail(msg: str) -> None:
    print(f"\n✗ HỎNG: {msg}", flush=True)
    sys.exit(1)


def _request(path: str, *, data=None, headers=None, method="GET"):
    url = f"{BASE_URL}{path}"
    if API_KEY:
        sep = "&" if "?" in url else "?"
        url = f"{url}{sep}api_key={API_KEY}"
    req = urllib.request.Request(url, data=data, method=method)
    if API_KEY:
        req.add_header("X-API-Key", API_KEY)
    for k, v in (headers or {}).items():
        req.add_header(k, v)
    return urllib.request.urlopen(req, timeout=30)


def build_drawing() -> tuple[bytes, str]:
    """Bản vẽ .dxf thật, có một tuyến ống ở bản vẽ chính. Trả về (nội dung, tên file).

    KHÁC với tầng CI: không dựng XREF ở đây. XREF cần file thứ hai nằm cạnh file chính
    trên đĩa của server, mà API chỉ nhận một file — kiểm phần đó phải đặt sẵn file lên
    server, xem `docs/E2E.md`.
    """
    import ezdxf

    doc = ezdxf.new()
    doc.units = ezdxf.units.MM
    msp = doc.modelspace()
    msp.add_line((0, 0), (MAIN_PIPE_LEN, 0), dxfattribs={"layer": "P-CHW-SUPPLY"})
    msp.add_line((0, 1000), (XREF_PIPE_LEN, 1000), dxfattribs={"layer": "P-CHW-SUPPLY"})

    buf = io.StringIO()
    doc.write(buf)
    return buf.getvalue().encode("utf-8"), "e2e_smoke.dxf"


def upload(content: bytes, filename: str) -> str:
    boundary = "----MEPAgentsE2E"
    body = b"".join([
        f"--{boundary}\r\n".encode(),
        f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'.encode(),
        b"Content-Type: application/dxf\r\n\r\n",
        content,
        f"\r\n--{boundary}--\r\n".encode(),
    ])
    headers = {"Content-Type": f"multipart/form-data; boundary={boundary}"}
    try:
        with _request("/api/v1/takeoff", data=body, headers=headers, method="POST") as resp:
            import json
            payload = json.load(resp)
    except urllib.error.HTTPError as e:
        _fail(f"upload trả về HTTP {e.code}: {e.read()[:300]!r}")
    except urllib.error.URLError as e:
        _fail(f"không kết nối được {BASE_URL} ({e.reason}). API đã chạy chưa?")
    task_id = payload.get("task_id")
    if not task_id:
        _fail(f"upload không trả về task_id: {payload}")
    return task_id


def wait_for_result(task_id: str) -> dict:
    import json

    deadline = time.time() + TIMEOUT_S
    last = None
    while time.time() < deadline:
        with _request(f"/api/v1/task/{task_id}") as resp:
            payload = json.load(resp)
        status = payload.get("status")
        if status != last:
            _step(f"trạng thái: {status}")
            last = status
        if status == "success":
            return payload
        if status == "error":
            _fail(f"worker báo lỗi: {payload.get('logs')}")
        time.sleep(2)
    _fail(
        f"quá {TIMEOUT_S:.0f}s vẫn chưa xong (trạng thái cuối: {last}).\n"
        "   Thường là worker KHÔNG nhặt được task: kiểm tra `docker compose logs worker` và\n"
        "   biến CELERY_BROKER_URL — trong container, 'localhost' là chính container đó."
    )


def download_and_check(task_id: str) -> None:
    with _request(f"/api/v1/download/{task_id}") as resp:
        data = resp.read()
        ctype = resp.headers.get("Content-Type", "")
    if not ctype.startswith("application/vnd.openxmlformats"):
        _fail(f"tải về không phải file Excel (Content-Type: {ctype}, {len(data)} byte): {data[:200]!r}")
    if len(data) < 1000:
        _fail(f"file Excel tải về chỉ {len(data)} byte — nhiều khả năng rỗng")

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    numbers = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)):
                    numbers.append(float(cell))
    if not any(abs(n - EXPECTED_M) <= EXPECTED_M * 0.02 for n in numbers):
        _fail(
            f"không thấy tổng chiều dài {EXPECTED_M} m trong Excel.\n"
            f"   Các số đọc được: {sorted(set(numbers))[:40]}"
        )
    _step(f"Excel hợp lệ ({len(data)} byte), tổng chiều dài khớp {EXPECTED_M} m")


def main() -> None:
    print(f"E2E smoke → {BASE_URL} (khóa API: {'có' if API_KEY else 'không'})\n")

    _step("kiểm tra API sống")
    try:
        with _request("/") as resp:
            resp.read()
    except urllib.error.URLError as e:
        _fail(f"không kết nối được {BASE_URL} ({e.reason}). Đã `docker compose up` chưa?")

    _step("dựng bản vẽ .dxf")
    content, filename = build_drawing()

    _step(f"tải lên {filename} ({len(content)} byte)")
    task_id = upload(content, filename)
    _step(f"task_id = {task_id}")

    _step("chờ worker xử lý")
    wait_for_result(task_id)

    _step("tải file BOQ về và kiểm nội dung")
    download_and_check(task_id)

    print("\n✓ ĐẠT — trọn đường: tải lên → worker → Excel → tải về")


if __name__ == "__main__":
    main()
