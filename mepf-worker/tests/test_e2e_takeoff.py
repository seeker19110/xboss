"""Kịch bản E2E tối thiểu: bản vẽ .dxf → API → worker → Excel BOQ → tải về.

**Vì sao cần:** toàn bộ test còn lại là unit/integration ở mức module, mock Celery/Redis
và không bao giờ ghi ra file Excel thật. Lớp lỗi mà chúng không bắt được đã xảy ra thật:
XREF bị loại khỏi khối lượng (PR #32) chỉ lộ ra khi chạy trọn đường từ bản vẽ tới con số.

**Phạm vi thật của file này — đọc kỹ trước khi tin:**

| Thành phần | Thật hay giả |
|---|---|
| File `.dxf` đầu vào | THẬT — dựng bằng ezdxf, có hình học và XREF thật |
| Bóc khối lượng (`auto_quantity_takeoff`) | THẬT — hình học, không mock |
| File Excel đầu ra | THẬT — ghi ra đĩa, đọc lại bằng thư viện đọc Excel |
| Endpoint FastAPI | THẬT — qua `TestClient` |
| Broker Celery + Redis | **GIẢ** — chạy đồng bộ trong tiến trình |

Nói cách khác: file này bao được "bản vẽ → khối lượng → Excel → tải về", **không** bao
được "worker rời thật sự nhận task qua Redis". Phần đó cần hạ tầng thật, xem
`scripts/e2e_smoke.py` và `docs/E2E.md`.
"""
from __future__ import annotations

import math
import os

import ezdxf
import pytest

# Chiều dài ống dựng sẵn trong bản vẽ, dùng để đối chiếu con số cuối cùng.
MAIN_PIPE_LEN = 12_000.0   # mm — nằm ở bản vẽ chính
XREF_PIPE_LEN = 8_000.0    # mm — nằm trong file XREF đi kèm

#: `auto_quantity_takeoff` cộng thêm hao hụt vật tư theo định mức vào chiều dài ống.
WASTE_FACTOR = 1.05


def _expected_m(*lengths_mm: float) -> float:
    return sum(lengths_mm) / 1000.0 * WASTE_FACTOR


@pytest.fixture
def drawing(tmp_path, monkeypatch):
    """Bản vẽ .dxf thật: một tuyến ống ở bản vẽ chính + một tuyến nằm trong XREF."""
    from src.workspace import set_workspace_dir

    set_workspace_dir(str(tmp_path))
    monkeypatch.chdir(tmp_path)

    # Đặt đơn vị bản vẽ là MILIMET. Mặc định của `ezdxf.new()` là MÉT, khi đó tọa độ
    # 12000 bị hiểu là 12000 m — đúng theo header nhưng không phải ý định của bài test.
    xref = ezdxf.new()
    xref.units = ezdxf.units.MM
    xref.modelspace().add_line((0, 0), (XREF_PIPE_LEN, 0), dxfattribs={"layer": "P-CHW-SUPPLY"})
    xref.saveas(str(tmp_path / "phu.dxf"))

    doc = ezdxf.new()
    doc.units = ezdxf.units.MM
    msp = doc.modelspace()
    msp.add_line((0, 0), (MAIN_PIPE_LEN, 0), dxfattribs={"layer": "P-CHW-SUPPLY"})
    doc.blocks.new("REF1", dxfattribs={"flags": 4, "xref_path": "phu.dxf"})
    msp.add_blockref("REF1", (0, 5_000))
    path = tmp_path / "tang1.dxf"
    doc.saveas(str(path))
    return path


def _read_excel_text(path: str) -> str:
    """Toàn bộ ô trong file Excel gộp thành một chuỗi, để soi nội dung mà không phụ thuộc
    vào cấu trúc cột cụ thể (cấu trúc BOQ còn thay đổi theo mẫu hồ sơ)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells.extend(str(c) for c in row if c is not None)
    return "\n".join(cells)


def _numbers_in(text: str) -> list[float]:
    out = []
    for token in text.replace(",", "").split():
        try:
            out.append(float(token))
        except ValueError:
            continue
    return out


def test_e2e_drawing_to_boq_excel(drawing, tmp_path):
    """Đường xương sống: bản vẽ thật → bóc khối lượng thật → Excel thật, và con số trong
    Excel phải khớp hình học đã dựng."""
    from src.tools import auto_quantity_takeoff

    out_path = str(tmp_path / "boq.xlsx")
    report = auto_quantity_takeoff.invoke({
        "file_path": "tang1.dxf",
        "output_excel_path": "boq.xlsx",
    })

    assert os.path.exists(out_path), f"không sinh ra file Excel. Báo cáo:\n{report}"

    text = _read_excel_text(out_path)
    numbers = _numbers_in(text)
    total_m = _expected_m(MAIN_PIPE_LEN, XREF_PIPE_LEN)

    assert any(math.isclose(n, total_m, rel_tol=0.02) for n in numbers), (
        f"không thấy tổng chiều dài {total_m} m trong Excel.\n"
        f"Các số đọc được: {sorted(set(numbers))[:40]}\nBáo cáo:\n{report}"
    )


def test_e2e_xref_content_is_counted(drawing, tmp_path):
    """Hồi quy PR #32: nội dung trong XREF từng bị loại khỏi khối lượng trong im lặng.
    Đây chính là loại lỗi mà chỉ chạy trọn đường mới bắt được."""
    from src.tools import auto_quantity_takeoff

    report = auto_quantity_takeoff.invoke({
        "file_path": "tang1.dxf",
        "output_excel_path": "boq.xlsx",
    })
    numbers = _numbers_in(_read_excel_text(str(tmp_path / "boq.xlsx")))
    main_only_m = _expected_m(MAIN_PIPE_LEN)
    total_m = _expected_m(MAIN_PIPE_LEN, XREF_PIPE_LEN)

    assert any(math.isclose(n, total_m, rel_tol=0.02) for n in numbers), (
        f"không thấy tổng {total_m} m (gồm cả XREF) trong Excel.\n"
        f"Các số đọc được: {sorted(set(numbers))[:40]}\nBáo cáo:\n{report}"
    )
    assert not any(math.isclose(n, main_only_m, rel_tol=0.001) for n in numbers), (
        f"có dòng bằng đúng phần bản vẽ chính ({main_only_m} m) — dấu hiệu XREF bị bỏ sót.\n"
        f"Báo cáo:\n{report}"
    )
    assert "phu.dxf" in report, f"không nêu việc đã gộp XREF vào kết quả:\n{report}"


def test_e2e_missing_xref_is_reported_not_silently_skipped(tmp_path, monkeypatch):
    """Nguyên tắc số 2 của dự án: thiếu dữ liệu phải nói rõ. Bản vẽ tham chiếu tới file
    XREF không có thật thì kết quả phải nêu tên file thiếu, chứ không được trả về một con
    số khối lượng thiếu mà người đọc tưởng là đủ."""
    from src.tools import auto_quantity_takeoff
    from src.workspace import set_workspace_dir

    set_workspace_dir(str(tmp_path))
    monkeypatch.chdir(tmp_path)

    doc = ezdxf.new()
    doc.units = ezdxf.units.MM
    doc.modelspace().add_line((0, 0), (MAIN_PIPE_LEN, 0), dxfattribs={"layer": "P-CHW-SUPPLY"})
    doc.blocks.new("REF1", dxfattribs={"flags": 4, "xref_path": "khong_ton_tai.dxf"})
    doc.modelspace().add_blockref("REF1", (0, 5_000))
    doc.saveas(str(tmp_path / "thieu_xref.dxf"))

    report = auto_quantity_takeoff.invoke({
        "file_path": "thieu_xref.dxf",
        "output_excel_path": "boq2.xlsx",
    })

    assert "khong_ton_tai.dxf" in report, f"không nêu tên file XREF thiếu:\n{report}"


def test_e2e_api_upload_to_download(drawing, tmp_path, monkeypatch):
    """Đường đi của Web App: POST /api/v1/takeoff → GET /api/v1/task/{id} →
    GET /api/v1/download/{id}, chạy qua FastAPI thật.

    Broker được thay bằng chạy đồng bộ trong tiến trình — đây là ranh giới của tầng test
    này, phần worker rời thật sự nằm ở `scripts/e2e_smoke.py`.
    """
    from fastapi.testclient import TestClient

    import src.api as api
    import src.celery_app as celery_mod

    results: dict[str, dict] = {}

    class _SyncTask:
        """Chạy thẳng thân task, không qua Redis."""

        def __init__(self, task_id, payload):
            self.id = task_id
            results[task_id] = payload

    def fake_delay(dwg_path, user_id="web_client"):
        task_id = f"e2e-{os.path.basename(dwg_path)}"
        out = os.path.join("data", "boq", f"boq_{os.path.basename(dwg_path)}.xlsx")
        os.makedirs(os.path.dirname(out), exist_ok=True)
        from src.tools import auto_quantity_takeoff
        logs = auto_quantity_takeoff.invoke({"file_path": dwg_path, "output_excel_path": out})
        return _SyncTask(task_id, {"status": "success", "excel_path": out, "logs": logs})

    monkeypatch.setattr(api.parse_cad_to_db_task, "delay", fake_delay)

    class _FakeAsyncResult:
        def __init__(self, task_id, app=None):
            self.id = task_id
            self._payload = results.get(task_id)

        @property
        def state(self):
            return "SUCCESS" if self._payload else "PENDING"

        @property
        def result(self):
            return self._payload

    monkeypatch.setattr(api, "AsyncResult", _FakeAsyncResult)
    monkeypatch.setattr(api, "UPLOAD_DIR", str(tmp_path))
    monkeypatch.setattr(celery_mod, "parse_cad_to_db_task", api.parse_cad_to_db_task, raising=False)

    client = TestClient(api.app)

    with open(drawing, "rb") as f:
        resp = client.post("/api/v1/takeoff", files={"file": ("tang1.dxf", f, "application/dxf")})
    assert resp.status_code == 200, resp.text
    task_id = resp.json()["task_id"]

    status = client.get(f"/api/v1/task/{task_id}")
    assert status.status_code == 200, status.text
    assert status.json()["status"] == "success", status.json()

    download = client.get(f"/api/v1/download/{task_id}")
    assert download.status_code == 200, download.text
    assert download.headers["content-type"].startswith(
        "application/vnd.openxmlformats"
    ), download.headers
    assert len(download.content) > 1000, "file Excel tải về rỗng bất thường"
